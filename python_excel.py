from openpyxl import load_workbook


# 打开excel
wb = load_workbook("d:/桌面/test.xlsx")

# 定位表单
sheet = wb['Sheet1']

# 定位单元格行列值
res = sheet.cell(1, 1).value

print("最大行：{0}".format(sheet.max_row))
print("最大列：{0}".format(sheet.max_column))

print('取值为：', res)

