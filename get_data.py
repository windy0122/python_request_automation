from openpyxl import load_workbook


class GetData(object):
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        # 打开excel
        wb = load_workbook(self.file_name)

        # 定位表单
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row+1):
            sub_data = dict()
            sub_data['login_code'] = sheet.cell(i, 1).value
            sub_data['password'] = sheet.cell(i, 2).value
            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    date = GetData('test.xlsx', 'Sheet1')
    print(date.get_data())
